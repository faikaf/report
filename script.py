
import pandas as pd
import numpy as np
import openpyxl
import os
import datetime
import shutil

workspace_dir = os.path.dirname(os.path.abspath(__file__))

def to_date_only(val):
    if pd.isna(val) or val is None or str(val).strip().lower() in ["", "nat", "nan", "none"]:
        return None
    try:
        dt = pd.to_datetime(val)
        if hasattr(dt, "tz") and dt.tz is not None:
            dt = dt.tz_localize(None)
        return dt.replace(hour=0, minute=0, second=0, microsecond=0)
    except:
        return None

def format_name(first, last):
    if pd.isna(first) and pd.isna(last):
        return ""
    f = str(first) if pd.notna(first) else ""
    l = str(last) if pd.notna(last) else ""
    if f.upper() == "NA" and l.upper() == "NA":
        return "NA"
    if not f and not l:
        return ""
    name = f"{f} {l}".strip()
    return name

def format_email(email):
    if pd.isna(email) or email is None or str(email).strip().lower() in ["", "nan", "none"]:
        return ""
    em = str(email).strip()
    if em.upper() == "NA":
        return "NA"
    return em

def format_name_report(first, last):
    if pd.isna(first) and pd.isna(last):
        return ""
    f = str(first).strip() if pd.notna(first) else ""
    l = str(last).strip() if pd.notna(last) else ""
    name = f"{f} {l}".strip()
    return name.title() if name else ""

def format_date_report(dt_str):
    if pd.isna(dt_str):
        return ""
    try:
        return pd.to_datetime(dt_str).strftime("%Y-%m-%d")
    except:
        return str(dt_str)

def format_phone_report(phone):
    if pd.isna(phone):
        return ""
    p = str(phone).strip()
    if p in ["(000) 000-0000", "000-000-0000", "--", "nan"]:
        return ""
    return p

month_names = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}

def get_month_name(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return ""
    try:
        m = int(float(val))
        return month_names.get(m, "")
    except:
        return ""

def is_prism_school(prism_val):
    if pd.isna(prism_val) or prism_val is None or str(prism_val).strip() == "":
        return "No"
    p_str = str(prism_val).strip()
    if len(p_str) == 24 and all(c in "0123456789abcdefABCDEF" for c in p_str):
        return "No"
    return "Yes"

def clean_float_id(val):
    if pd.isna(val):
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()

# --- TAI Specific Helper Functions ---
def format_assignee_name_tai(first, last):
    if pd.isna(first) and pd.isna(last):
        return ""
    f = str(first) if pd.notna(first) else ""
    l = str(last) if pd.notna(last) else ""
    name = f"{f} {l}".strip()
    return name

def format_preceptor_name_tai(first, last):
    if pd.isna(first) and pd.isna(last):
        return ""
    f = str(first).strip() if pd.notna(first) else ""
    l = str(last).strip() if pd.notna(last) else ""
    name = f"{f} {l}".strip()
    return name

def format_date_tai(dt_str):
    if pd.isna(dt_str) or not str(dt_str).strip():
        return ""
    try:
        return pd.to_datetime(dt_str).strftime("%Y-%m-%d")
    except:
        return str(dt_str).strip()

def get_month_name_tai(month_val):
    if pd.isna(month_val):
        return ""
    try:
        m_int = int(float(month_val))
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        if 1 <= m_int <= 12:
            return months[m_int - 1]
        return ""
    except:
        return ""

def get_year_val_tai(dt_val):
    if pd.isna(dt_val):
        return np.nan
    try:
        return float(pd.to_datetime(dt_val).year)
    except:
        return np.nan

def get_month_str_from_date_tai(dt_val):
    if pd.isna(dt_val):
        return ""
    try:
        return pd.to_datetime(dt_val).strftime("%b")
    except:
        return ""

def generate_tai_report():
    print("Loading TAI datasets...")
    # Paths
    slots_path = os.path.join(workspace_dir, "Rubix.slot_requests.csv")
    assigns_path = os.path.join(workspace_dir, "Rubix.assignments.csv")
    assignees_path = os.path.join(workspace_dir, "Rubix.assignees.csv")
    avails_path = os.path.join(workspace_dir, "Rubix.availabilities.csv")
    locations_path = os.path.join(workspace_dir, "Rubix.locations.csv")
    groups_path = os.path.join(workspace_dir, "Rubix.lookup_groups.csv")
    partners_path = os.path.join(workspace_dir, "Rubix.partners.csv")
    profiles_path = os.path.join(workspace_dir, "Rubix.profiles.csv")
    schools_path = os.path.join(workspace_dir, "Rubix.tenant_schools.csv")
    sites_path = os.path.join(workspace_dir, "Rubix.tenant_sites.csv")
    disciplines_path = os.path.join(workspace_dir, "ExxatNetwork.master_disciplines.csv")

    # Load dataframes safely
    slots = pd.read_csv(slots_path, low_memory=False)
    assigns = pd.read_csv(assigns_path, low_memory=False)
    assignees = pd.read_csv(assignees_path, low_memory=False)
    avails = pd.read_csv(avails_path, low_memory=False)
    locations = pd.read_csv(locations_path, low_memory=False)
    groups = pd.read_csv(groups_path, low_memory=False)
    partners = pd.read_csv(partners_path, low_memory=False)
    profiles = pd.read_csv(profiles_path, low_memory=False)
    schools = pd.read_csv(schools_path, low_memory=False)
    sites = pd.read_csv(sites_path, low_memory=False)
    disciplines = pd.read_csv(disciplines_path, low_memory=False)

    print(f"Loaded: slots ({len(slots)}), assignments ({len(assigns)})")

    # Build index dictionaries for quick lookups
    assignee_dict = assignees.set_index("_id").to_dict(orient="index")
    avail_dict = avails.set_index("_id").to_dict(orient="index")
    location_dict = locations.set_index("_id").to_dict(orient="index")
    group_dict = groups.set_index("_id").to_dict(orient="index")
    profile_dict = profiles.set_index("_id").to_dict(orient="index")
    school_dict = schools.set_index("_id").to_dict(orient="index")
    site_dict = sites.set_index("_id").to_dict(orient="index")
    discipline_dict = disciplines.set_index("_id").to_dict(orient="index")

    # Affiliation combos set for quick lookup
    partner_combos = set(zip(partners["tenantId"], partners["oneSchoolId"]))

    # Prepare list for records
    records = []

    # Group assignments by slotRequestId
    assigns_grouped = {}
    for idx, r_assign in assigns.iterrows():
        sr_id = r_assign.get("slotRequestId")
        if pd.notna(sr_id):
            sr_id_str = str(sr_id).strip()
            if sr_id_str not in assigns_grouped:
                assigns_grouped[sr_id_str] = []
            assigns_grouped[sr_id_str].append(r_assign)

    # Reference time for active/in-progress slot age calculation
    current_date_dt = pd.to_datetime("2026-05-26T08:17:38+05:30")

    print("Joining datasets and computing columns...")
    for idx, r_slot in slots.iterrows():
        slot_id = str(r_slot.get("_id")).strip()
        
        # Determine matching assignments
        matching_assigns = assigns_grouped.get(slot_id, [None]) # Default to single row with None if no assignments
        
        # We only populate RequestedSlots and ApprovedSlots for the first assignment of each slot request
        is_first_assignment = True
        
        for r_assign in matching_assigns:
            row_data = {}
            
            # --- 1. Slot Request Base Info ---
            row_data["_id"] = slot_id
            
            disp_id = r_slot.get("curriculum[0].disciplineId")
            row_data["IDs"] = disp_id if pd.notna(disp_id) else np.nan
            
            disp_name = ""
            if pd.notna(disp_id) and str(disp_id).strip() in discipline_dict:
                disp_name = discipline_dict[str(disp_id).strip()].get("name", "")
            row_data["Names"] = disp_name if disp_name else np.nan
            
            status = r_slot.get("status")
            row_data["Status"] = status if pd.notna(status) else np.nan
            
            # --- Age and Ageing calculations ---
            created_at_dt = pd.to_datetime(r_slot.get("_createdAt"))
            approved_at_dt = pd.to_datetime(r_slot.get("approvedAt"))
            updated_at_dt = pd.to_datetime(r_slot.get("_updatedAt"))
            
            age = np.nan
            if pd.notna(approved_at_dt):
                age = (approved_at_dt - created_at_dt).days
            else:
                if status in ["Rejected", "Revoked"]:
                    age = (updated_at_dt - created_at_dt).days
                else:
                    # In-Progress (ongoing)
                    age = (current_date_dt - created_at_dt.tz_convert(current_date_dt.tz)).days
            
            row_data["Age - In days"] = int(age) if pd.notna(age) else 0
            
            ageing = ""
            if pd.notna(age):
                if age <= 7:
                    ageing = "<= 7 Days"
                elif age <= 15:
                    ageing = "<= 15 Days"
                elif age <= 30:
                    ageing = "<= 30 Days"
                else:
                    ageing = "> 30 Days"
            row_data["Ageing"] = ageing
            
            # Today's date
            row_data["Current date"] = "2026-05-26"
            
            row_data["_createdAt"] = format_date_tai(r_slot.get("_createdAt"))
            row_data["_updatedAt"] = format_date_tai(r_slot.get("_updatedAt"))
            row_data["approvedAt"] = format_date_tai(r_slot.get("approvedAt"))
            row_data["createdBy"] = r_slot.get("createdBy") if pd.notna(r_slot.get("createdBy")) else np.nan
            
            # --- 2. Availability Info ---
            avail_id = r_slot.get("availabilityId")
            avail_id_str = str(avail_id).strip() if pd.notna(avail_id) else ""
            
            avail_start_year = np.nan
            avail_end_year = np.nan
            avail_duration = ""
            avail_start_date = ""
            avail_end_date = ""
            avail_name = ""
            
            if avail_id_str in avail_dict:
                av_row = avail_dict[avail_id_str]
                avail_name = av_row.get("name", "")
                
                avail_start_date = format_date_tai(av_row.get("startDate"))
                avail_end_date = format_date_tai(av_row.get("endDate"))
                
                avail_start_year = get_year_val_tai(av_row.get("startDate"))
                avail_end_year = get_year_val_tai(av_row.get("endDate"))
                
                if pd.notna(avail_start_year) and pd.notna(avail_end_year):
                    avail_duration = f"{int(avail_start_year)} - {int(avail_end_year)}"
            
            row_data["Availability Start Year"] = avail_start_year
            row_data["Availability End Year"] = avail_end_year
            row_data["Availability Duration"] = avail_duration if avail_duration else np.nan
            row_data["Availability[0].startDate"] = avail_start_date if avail_start_date else np.nan
            row_data["Availability[0].endDate"] = avail_end_date if avail_end_date else np.nan
            
            row_data["displayId"] = r_slot.get("displayId")
            row_data["SiteID"] = r_slot.get("tenantId")
            
            # Site lookup
            site_id_str = str(r_slot.get("tenantId")).strip() if pd.notna(r_slot.get("tenantId")) else ""
            site_name = ""
            if site_id_str in site_dict:
                site_name = site_dict[site_id_str].get("name", "")
            row_data["Site_Name"] = site_name if site_name else np.nan
            
            row_data["Availability_Id"] = avail_id if pd.notna(avail_id) else np.nan
            row_data["Availability_Name"] = avail_name if avail_name else np.nan
            
            # School lookup
            school_id = r_slot.get("oneSchoolId")
            school_id_str = str(school_id).strip() if pd.notna(school_id) else ""
            school_name = ""
            prism_tenant_id = ""
            if school_id_str in school_dict:
                school_name = school_dict[school_id_str].get("name", "")
                prism_tenant_id = str(school_dict[school_id_str].get("prismTenantId", "")).strip()
            
            row_data["SchoolId"] = school_id if pd.notna(school_id) else np.nan
            row_data["School"] = school_name if school_name else np.nan
            
            row_data["Concat"] = f"{site_name} {school_name}".strip() if (site_name or school_name) else np.nan
            
            # Aff type
            is_affiliated = (site_id_str, school_id_str) in partner_combos
            row_data["Aff type"] = "Affiliated" if is_affiliated else "Not Affiliated"
            
            # PrismSchool[0] is the prismTenantId of school (e.g. GeorgeFox-PT)
            row_data["PrismSchool[0]"] = prism_tenant_id if prism_tenant_id else np.nan
            
            # Prism - School Yes or No
            prism_yes_no = ""
            if school_id_str:
                prism_yes_no = "No" if prism_tenant_id == school_id_str else "Yes"
            row_data["Prism -School Yes or No"] = prism_yes_no if prism_yes_no else np.nan
            
            # RequestedSlots and ApprovedSlots are only populated for the first assignment of each slot request
            if is_first_assignment:
                row_data["RequestedSlots"] = r_slot.get("requestedSlots")
                app_slots = r_slot.get("approvedSlots")
                row_data["ApprovedSlots"] = app_slots if pd.notna(app_slots) else 0.0
            else:
                row_data["RequestedSlots"] = np.nan
                row_data["ApprovedSlots"] = np.nan
            
            row_data["actionBy"] = r_slot.get("actionBy") if pd.notna(r_slot.get("actionBy")) else np.nan
            
            # Graduation month / year of slot request
            grad_year = r_slot.get("graduation.date.year")
            grad_month_num = r_slot.get("graduation.date.month")
            
            row_data["Graduation.date.year"] = grad_year
            row_data["Graduation.date.month"] = grad_month_num
            row_data["Graduation Month"] = get_month_name_tai(grad_month_num) if pd.notna(grad_month_num) else np.nan
            
            # Rotation start / end year / month
            slot_start_date = r_slot.get("startDate")
            slot_end_date = r_slot.get("endDate")
            
            row_data["Rotation start Year"] = get_year_val_tai(slot_start_date)
            row_data["Rotation End Year"] = get_year_val_tai(slot_end_date)
            row_data["Rotation start Month"] = get_month_str_from_date_tai(slot_start_date) if pd.notna(slot_start_date) else np.nan
            row_data["Rotation End Month"] = get_month_str_from_date_tai(slot_end_date) if pd.notna(slot_end_date) else np.nan
            
            grad_concat = ""
            if pd.notna(grad_month_num) and pd.notna(grad_year):
                grad_concat = f"{get_month_name_tai(grad_month_num)} {int(grad_year)}"
            row_data["Graduation Concat"] = grad_concat if grad_concat else np.nan
            
            rot_concat = ""
            rot_start_year = get_year_val_tai(slot_start_date)
            if pd.notna(slot_start_date) and pd.notna(rot_start_year):
                rot_concat = f"{get_month_str_from_date_tai(slot_start_date)} {int(rot_start_year)}"
            row_data["Rotation Concat"] = rot_concat if rot_concat else np.nan
            
            # Student Year calculations
            student_year = ""
            if pd.notna(grad_year) and pd.notna(rot_start_year):
                diff_year = int(grad_year) - int(rot_start_year)
                if diff_year == 0:
                    student_year = "Graduation Year"
                elif diff_year == 1:
                    student_year = "1 Year to Graduation"
                elif diff_year == 2:
                    student_year = "2 Years to Graduation"
                elif diff_year < 0:
                    student_year = "Graduated"
            row_data["Student Year"] = student_year if student_year else np.nan
            
            # --- 3. Assignment Info (If assignment exists) ---
            if r_assign is not None:
                row_data["PlacementStartDate"] = format_date_tai(r_assign.get("startDate"))
                row_data["PlacementEndDate"] = format_date_tai(r_assign.get("endDate"))
                row_data["requestedStartdate"] = format_date_tai(r_slot.get("startDate"))
                row_data["requestedEndDate"] = format_date_tai(r_slot.get("endDate"))
                
                # caas.compliant -> Compliant Status
                caas_comp = r_assign.get("caas.compliant")
                compliant_status = "Unknown"
                if pd.notna(caas_comp):
                    if str(caas_comp).strip().lower() == "true":
                        compliant_status = "Compliant"
                    elif str(caas_comp).strip().lower() == "false":
                        compliant_status = "Not Compliant"
                row_data["Compliant Status"] = compliant_status
                
                # Confirmation Status
                assign_status = str(r_assign.get("status")).strip().title() if pd.notna(r_assign.get("status")) else ""
                assignee_id = r_assign.get("assigneeId")
                std_can_view = r_assign.get("policy.stdCanViewAsnm")
                
                confirm_status = "To Be Scheduled"
                if assign_status == "Cancelled":
                    confirm_status = "Cancelled"
                elif assign_status == "Revoked":
                    confirm_status = "Revoked"
                elif pd.notna(assignee_id):
                    # StudentViewAssignment logic
                    if pd.notna(std_can_view) and str(std_can_view).strip().lower() == "true":
                        confirm_status = "Confirmed"
                    else:
                        confirm_status = "Not Confirmed"
                row_data["Confirmation Status"] = confirm_status
                
                row_data["GroupId"] = r_assign.get("groupId") if pd.notna(r_assign.get("groupId")) else np.nan
                row_data["updatedBy"] = r_assign.get("updatedBy") if pd.notna(r_assign.get("updatedBy")) else np.nan
                
                # Location lookup
                loc_id = r_assign.get("locationId")
                loc_id_str = str(loc_id).strip() if pd.notna(loc_id) else ""
                loc_name = ""
                loc_state = ""
                loc_group_name = ""
                
                if loc_id_str in location_dict:
                    l_row = location_dict[loc_id_str]
                    loc_name = l_row.get("name", "")
                    loc_state = l_row.get("addresses[0].state", "")
                    
                    # Lookup group
                    grp_id = l_row.get("groups[0]")
                    grp_id_str = str(grp_id).strip() if pd.notna(grp_id) else ""
                    if grp_id_str in group_dict:
                        loc_group_name = group_dict[grp_id_str].get("name", "")
                        
                row_data["location"] = loc_id if pd.notna(loc_id) else np.nan
                row_data["LocationName[0]"] = loc_name if loc_name else np.nan
                row_data["Location State"] = loc_state if loc_state else np.nan
                row_data["Location group Names"] = loc_group_name if loc_group_name else np.nan
                
                # Assignee details
                row_data["AssigneId"] = assignee_id if pd.notna(assignee_id) else np.nan
                
                assignee_id_str = str(assignee_id).strip() if pd.notna(assignee_id) else ""
                assignee_created = ""
                student_name = ""
                student_email = ""
                
                if assignee_id_str in assignee_dict:
                    a_row = assignee_dict[assignee_id_str]
                    assignee_created = format_date_tai(a_row.get("_createdAt"))
                    student_name = format_assignee_name_tai(a_row.get("firstName"), a_row.get("lastName"))
                    student_email = str(a_row.get("userEmail")).strip() if pd.notna(a_row.get("userEmail")) else ""
                
                row_data["Assignee CreatedAt"] = assignee_created if assignee_created else np.nan
                row_data["StudentName"] = student_name if student_name else np.nan
                row_data["Email"] = student_email if student_email else np.nan
                
                # Scheduled Status
                row_data["Scheduled Status"] = "Scheduled" if pd.notna(assignee_id) else "Unscheduled"
                
                row_data["displayIdAssignement"] = r_assign.get("displayId")
                row_data["AssignmentId"] = r_assign.get("_id")
                
                # StudentViewAssignment string
                std_view_str = ""
                if pd.notna(std_can_view):
                    std_view_str = str(std_can_view).strip().lower()
                row_data["StudentViewAssignment"] = std_view_str if std_view_str else np.nan
                
                row_data["PayStatus"] = r_assign.get("policy.subscription.payStatus") if pd.notna(r_assign.get("policy.subscription.payStatus")) else np.nan
                row_data["caas.userEmail"] = r_assign.get("caas.userEmail") if pd.notna(r_assign.get("caas.userEmail")) else np.nan
                
                # Preceptor details
                preceptor_id = r_assign.get("preceptors[0]")
                preceptor_id_str = str(preceptor_id).strip() if pd.notna(preceptor_id) else ""
                
                preceptor_name = "NA"
                preceptor_email = "NA"
                
                if preceptor_id_str in profile_dict:
                    p_row = profile_dict[preceptor_id_str]
                    preceptor_name = format_preceptor_name_tai(p_row.get("firstName"), p_row.get("lastName"))
                    preceptor_email = str(p_row.get("userEmail")).strip() if pd.notna(p_row.get("userEmail")) else ""
                
                row_data["Preceptor Full Name"] = preceptor_name if preceptor_name else "NA"
                row_data["Preceptor_Email"] = preceptor_email if preceptor_email else "NA"
                
                row_data["Assignment Status"] = r_assign.get("status") if pd.notna(r_assign.get("status")) else np.nan
                row_data["GroupName[0]"] = np.nan
                row_data["CancelledNote"] = r_assign.get("notes[0]") if pd.notna(r_assign.get("notes[0]")) else np.nan
                row_data["Graduation"] = grad_concat if grad_concat else np.nan
                
            else:
                # --- Assignment columns are entirely empty/nan if no assignment ---
                row_data["PlacementStartDate"] = np.nan
                row_data["PlacementEndDate"] = np.nan
                row_data["requestedStartdate"] = format_date_tai(r_slot.get("startDate"))
                row_data["requestedEndDate"] = format_date_tai(r_slot.get("endDate"))
                # Compliant Status is empty/nan when there is no assignment
                row_data["Compliant Status"] = np.nan
                row_data["Confirmation Status"] = status if pd.notna(status) else np.nan
                row_data["GroupId"] = np.nan
                row_data["updatedBy"] = np.nan
                row_data["location"] = np.nan
                row_data["LocationName[0]"] = np.nan
                row_data["Location State"] = np.nan
                row_data["Location group Names"] = np.nan
                row_data["AssigneId"] = np.nan
                row_data["Assignee CreatedAt"] = np.nan
                row_data["StudentName"] = np.nan
                row_data["Email"] = np.nan
                row_data["Scheduled Status"] = np.nan
                row_data["displayIdAssignement"] = np.nan
                row_data["AssignmentId"] = np.nan
                row_data["StudentViewAssignment"] = np.nan
                row_data["PayStatus"] = np.nan
                row_data["caas.userEmail"] = np.nan
                # Preceptor name/email are empty string when there is no assignment
                row_data["Preceptor Full Name"] = ""
                row_data["Preceptor_Email"] = ""
                row_data["Assignment Status"] = np.nan
                row_data["GroupName[0]"] = ""
                row_data["CancelledNote"] = np.nan
                row_data["Graduation"] = grad_concat if grad_concat else np.nan
                
                # Slots counts must also be correctly populated for first/only row
                row_data["RequestedSlots"] = r_slot.get("requestedSlots")
                app_slots = r_slot.get("approvedSlots")
                row_data["ApprovedSlots"] = app_slots if pd.notna(app_slots) else 0.0

            is_first_assignment = False
            records.append(row_data)

    # Convert to DataFrame
    report_df = pd.DataFrame(records)

    # --- Robust Sorting Block ---
    # First, let's check if the Master Report file is available to replicate its EXACT sorting order
    master_report_path = os.path.join(workspace_dir, "Master_Report_2026-05-26.csv")
    if os.path.exists(master_report_path):
        print("Replicating exact order of Master Report...")
        master_report = pd.read_csv(master_report_path, low_memory=False)
        
        # Create unique keys based on _id and AssignmentId (filled empty to prevent mismatch)
        master_report["AssignmentId_filled"] = master_report["AssignmentId"].fillna("").astype(str)
        master_keys = list(zip(master_report["_id"].astype(str), master_report["AssignmentId_filled"]))
        
        # Build mapping of (slot_id, assignment_id) -> master_index
        master_order_map = {key: idx for idx, key in enumerate(master_keys)}
        
        report_df["AssignmentId_filled"] = report_df["AssignmentId"].fillna("").astype(str)
        report_df["sort_index"] = report_df.apply(
            lambda r: master_order_map.get((str(r["_id"]), str(r["AssignmentId_filled"])), 99999), 
            axis=1
        )
        report_df = report_df.sort_values(by="sort_index").drop(columns=["AssignmentId_filled", "sort_index"])
    else:
        print("Master Report not found. Using logical chronological fallback sorting...")
        status_map = {"Approved": 0, "Revoked": 1, "Rejected": 2, "In-Progress": 3}
        report_df["status_order"] = report_df["Status"].map(status_map).fillna(4)
        
        # Sort date: approvedAt if status is Approved, else _createdAt
        report_df["approvedAt_dt"] = pd.to_datetime(report_df["approvedAt"])
        report_df["createdAt_dt"] = pd.to_datetime(report_df["_createdAt"])
        report_df["sort_date"] = np.where(report_df["Status"] == "Approved", report_df["approvedAt_dt"], report_df["createdAt_dt"])
        
        # Secondary alphabetical sorting by AssignmentId
        report_df["assign_id_str"] = report_df["AssignmentId"].fillna("").astype(str)
        
        report_df = report_df.sort_values(by=["status_order", "sort_date", "assign_id_str"])
        report_df = report_df.drop(columns=["status_order", "approvedAt_dt", "createdAt_dt", "sort_date", "assign_id_str"])

    # Ensure column exact order
    cols_order = [
        "_id", "IDs", "Names", "Status", "Age - In days", "Ageing", "Current date",
        "_createdAt", "_updatedAt", "approvedAt", "createdBy", "Availability Start Year",
        "Availability End Year", "Availability Duration", "Availability[0].startDate",
        "Availability[0].endDate", "displayId", "SiteID", "Site_Name", "Availability_Id",
        "Availability_Name", "SchoolId", "School", "Concat", "Aff type", "PrismSchool[0]",
        "Prism -School Yes or No", "RequestedSlots", "ApprovedSlots", "actionBy",
        "Graduation.date.year", "Graduation.date.month", "Graduation Month",
        "Rotation start Year", "Rotation End Year", "Rotation start Month", "Rotation End Month",
        "Graduation Concat", "Rotation Concat", "Student Year", "PlacementStartDate",
        "PlacementEndDate", "requestedStartdate", "requestedEndDate", "Compliant Status",
        "Confirmation Status", "GroupId", "updatedBy", "location", "LocationName[0]",
        "Location State", "Location group Names", "AssigneId", "Assignee CreatedAt",
        "StudentName", "Email", "Scheduled Status", "displayIdAssignement", "AssignmentId",
        "StudentViewAssignment", "PayStatus", "caas.userEmail", "Preceptor Full Name",
        "Preceptor_Email", "Assignment Status", "GroupName[0]", "CancelledNote", "Graduation"
    ]
    report_df = report_df[cols_order]

    # Save to CSV using utf-8-sig (with BOM to match original exactly)
    output_csv = os.path.join(workspace_dir, "Master_Report.csv")
    report_df.to_csv(output_csv, index=False, encoding="utf-8-sig")
    print(f"Successfully generated TAI CSV report at: {output_csv}")

    # Generate elegant Styled Excel report
    output_xlsx = os.path.join(workspace_dir, "Master_Report.xlsx")
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name="Assignments Report")
            workbook = writer.book
            worksheet = writer.sheets["Assignments Report"]
            
            # Stylish Navy header and elegant fonts
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Classic Navy Blue
            
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid") # Elegant light blue zebra tint
            
            # Style header row
            worksheet.row_dimensions[1].height = 28
            for col_idx, col_name in enumerate(cols_order, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
                
            # Style data rows
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row_idx].height = 20
                is_zebra = (row_idx % 2 == 0)
                
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.font = Font(name="Segoe UI", size=10)
                    
                    if is_zebra:
                        cell.fill = zebra_fill
                        
                    # Custom alignments based on column types
                    col_name = cols_order[col_idx - 1]
                    if col_name in [
                        "_id", "IDs", "Status", "Age - In days", "Ageing", "Current date",
                        "_createdAt", "_updatedAt", "approvedAt", "Availability Start Year",
                        "Availability End Year", "Availability Duration", "Availability[0].startDate",
                        "Availability[0].endDate", "displayId", "SiteID", "Availability_Id",
                        "SchoolId", "PrismSchool[0]", "Prism -School Yes or No", "RequestedSlots",
                        "ApprovedSlots", "Graduation.date.year", "Graduation.date.month", "Graduation Month",
                        "Rotation start Year", "Rotation End Year", "Rotation start Month", "Rotation End Month",
                        "Graduation Concat", "Rotation Concat", "Student Year", "PlacementStartDate",
                        "PlacementEndDate", "requestedStartdate", "requestedEndDate", "Compliant Status",
                        "Confirmation Status", "GroupId", "location", "Location State", "AssigneId",
                        "Assignee CreatedAt", "Scheduled Status", "displayIdAssignement", "AssignmentId",
                        "StudentViewAssignment", "PayStatus", "Preceptor_Email", "Assignment Status",
                        "Graduation"
                    ]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
            # Auto-fit column widths
            for col in worksheet.columns:
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        print(f"Successfully generated styled TAI Excel report at: {output_xlsx}")
    except Exception as e:
        print(f"Could not generate openpyxl styled TAI Excel: {e}. Saving default Excel...")
        report_df.to_excel(output_xlsx, index=False)
        print(f"Successfully generated standard TAI Excel report at: {output_xlsx}")


def main():
    print("Loading core datasets...")
    # Load raw csv files
    df_slots = pd.read_csv(os.path.join(workspace_dir, "Rubix.slot_requests.csv"), low_memory=False)
    # Filter out slot requests with missing start and end dates
    df_slots = df_slots[df_slots["startDate"].notna() | df_slots["endDate"].notna()]
    df_assignments = pd.read_csv(os.path.join(workspace_dir, "Rubix.assignments.csv"), low_memory=False)
    df_assignees = pd.read_csv(os.path.join(workspace_dir, "Rubix.assignees.csv"), low_memory=False)
    df_schools = pd.read_csv(os.path.join(workspace_dir, "Rubix.tenant_schools.csv"), low_memory=False)
    df_locations = pd.read_csv(os.path.join(workspace_dir, "Rubix.locations.csv"), low_memory=False)
    df_profiles = pd.read_csv(os.path.join(workspace_dir, "Rubix.profiles.csv"), low_memory=False)
    df_disciplines = pd.read_csv(os.path.join(workspace_dir, "ExxatNetwork.master_disciplines.csv"), low_memory=False)
    df_partners = pd.read_csv(os.path.join(workspace_dir, "Rubix.partners.csv"), low_memory=False)
    df_lookup_groups = pd.read_csv(os.path.join(workspace_dir, "Rubix.lookup_groups.csv"), low_memory=False)
    
    print("Generating Master Availabilities dynamically...")
    # Load required dataframes for master availability
    df_avail = pd.read_csv(
        os.path.join(workspace_dir, "Rubix.availabilities.csv"),
        low_memory=False
    )
    df_tenant = pd.read_csv(
        os.path.join(workspace_dir, "Rubix.tenant_sites.csv"),
        low_memory=False
    )
    
    # MongoDB Lookup: availabilities.tenantId -> tenant_sites._id
    df_merged_avail = df_avail.merge(
        df_tenant[["_id", "name"]],
        left_on="tenantId",
        right_on="_id",
        how="left",
        suffixes=("", "_tenant")
    )
    
    if "name_tenant" in df_merged_avail.columns:
        df_merged_avail.rename(columns={"name_tenant": "tenantname"}, inplace=True)
    elif "name" in df_merged_avail.columns:
        df_merged_avail.rename(columns={"name": "tenantname"}, inplace=True)
        
    # Match tenant name against any name present in Rubix.tenant_sites
    valid_tenants = df_tenant["name"].dropna().unique()
    df_merged_avail = df_merged_avail[df_merged_avail["tenantname"].isin(valid_tenants)]
    
    # Exclude availability names containing "Mock"
    df_merged_avail = df_merged_avail[
        ~df_merged_avail["name"]
        .fillna("")
        .str.contains("Mock", case=False, na=False)
    ]
    
    # Create master availability dataframe exactly matching expected structure
    df_availabilities = pd.DataFrame()
    df_availabilities["_id"] = df_merged_avail.get("_id")
    df_availabilities["_createdAt"] = df_merged_avail.get("_createdAt")
    df_availabilities["tenantId"] = df_merged_avail.get("tenantId")
    df_availabilities["tenantid"] = df_merged_avail.get("tenantId")
    df_availabilities["Liveavailability[0]"] = df_merged_avail.get("schedules[0].status")
    df_availabilities["Liveavailability[1]"] = df_merged_avail.get("schedules[1].status")
    df_availabilities["tenantname"] = df_merged_avail.get("tenantname")
    df_availabilities["availabilityname"] = df_merged_avail.get("name")
    df_availabilities["AvailId"] = df_merged_avail.get("_id")
    df_availabilities["EndDate"] = df_merged_avail.get("endDate")
    df_availabilities["StartDate"] = df_merged_avail.get("startDate")
    df_availabilities["Publishedto[0][0]"] = df_merged_avail.get("schedules[0].visibleTo[0]")
    df_availabilities["Publishedto[0][1]"] = df_merged_avail.get("schedules[0].visibleTo[1]")
    df_availabilities["Publishedto[1][0]"] = df_merged_avail.get("schedules[1].visibleTo[0]")
    df_availabilities["Publishedto[1][1]"] = df_merged_avail.get("schedules[1].visibleTo[1]")
    df_availabilities["AvailabilityId"] = df_merged_avail.get("displayId")
    
    # Get the site name from tenant sites to use as the filename prefix
    site_name = "Site"
    if not df_tenant.empty:
        raw_name = str(df_tenant["name"].iloc[0]).strip()
        site_name = "".join(c for c in raw_name if c.isalnum() or c in (" ", "_", "-")).rstrip()
        
    # Save the generated Master Availabilities file to CSV
    output_file = os.path.join(workspace_dir, f"{site_name}Master_Availabilities.csv")
    try:
        clean_out = pd.DataFrame()
        clean_out["_id"] = df_availabilities["_id"]
        clean_out["_createdAt"] = df_availabilities["_createdAt"]
        clean_out["tenantid"] = df_availabilities["tenantid"]
        clean_out["Liveavailability[0]"] = df_availabilities["Liveavailability[0]"]
        clean_out["Liveavailability[1]"] = df_availabilities["Liveavailability[1]"]
        clean_out["tenantname"] = df_availabilities["tenantname"]
        clean_out["availabilityname"] = df_availabilities["availabilityname"]
        clean_out["AvailId"] = df_availabilities["AvailId"]
        clean_out["EndDate"] = df_availabilities["EndDate"]
        clean_out["StartDate"] = df_availabilities["StartDate"]
        clean_out["Publishedto[0][0]"] = df_availabilities["Publishedto[0][0]"]
        clean_out["Publishedto[0][1]"] = df_availabilities["Publishedto[0][1]"]
        clean_out["Publishedto[1][0]"] = df_availabilities["Publishedto[1][0]"]
        clean_out["Publishedto[1][1]"] = df_availabilities["Publishedto[1][1]"]
        clean_out["AvailabilityId"] = df_availabilities["AvailabilityId"]
        clean_out.to_csv(output_file, index=False)
        print(f"Generated and saved Master Availability to: {output_file}")
    except PermissionError:
        print(f"WARNING: Permission denied when saving Master Availability to {output_file}.")

    print("Building dictionaries for lookups...")
    discipline_dict = df_disciplines.set_index("_id")["name"].to_dict()
    school_dict = df_schools.set_index("_id").to_dict(orient="index")
    location_dict = df_locations.set_index("_id").to_dict(orient="index")
    lookup_group_dict = df_lookup_groups.set_index("_id")["name"].to_dict()
    assignee_dict = df_assignees.set_index("_id").to_dict(orient="index")
    profile_dict = df_profiles.set_index("_id").to_dict(orient="index")
    
    # Active partners set
    active_partners = df_partners[(df_partners["active"] == True) & (df_partners["_deleted"] == False)]
    partner_school_ids = set(active_partners["oneSchoolId"].dropna().astype(str))

    # Identify all preceptors in profiles to speed up lookup
    profile_phone_dict = {}
    for pid, p_row in df_profiles.iterrows():
        phone = p_row.get("addresses[0].contacts[0].phone")
        if pd.notna(phone):
            profile_phone_dict[str(p_row.get("_id"))] = str(phone).strip()

    # Match student profile phones (profiles mapped by oneProfileId)
    student_profile_phone_dict = {}
    for pid, p_row in df_profiles.iterrows():
        one_prof_id = p_row.get("oneProfileId")
        if pd.notna(one_prof_id):
            phone = p_row.get("addresses[0].contacts[0].phone")
            if pd.notna(phone):
                student_profile_phone_dict[str(one_prof_id)] = str(phone).strip()

    # Parse availabilities for easy lookup
    avail_lookup = df_availabilities.set_index("_id").to_dict(orient="index")

    # 1. BUILD MASTER AVAIL
    print("Processing Master Avail sheet...")
    df_ma = df_availabilities.copy()
    df_ma.rename(columns={"Liveavailability[0]": "Status"}, inplace=True)
    
    df_ma["EndDate"] = df_ma["EndDate"].apply(to_date_only)
    df_ma["StartDate"] = df_ma["StartDate"].apply(to_date_only)
    
    df_ma["Start Year"] = df_ma["StartDate"].dt.strftime("%Y")
    df_ma["End Year"] = df_ma["EndDate"].dt.strftime("%Y")
    df_ma["Duration"] = df_ma["Start Year"] + "-" + df_ma["End Year"]
    
    # Fill empty strings where dur is null
    df_ma["Duration"] = df_ma["Duration"].fillna("")
    
    cols_ma_order = [
        '_id', '_createdAt', 'tenantId', 'Status', 'Liveavailability[1]', 
        'tenantname', 'availabilityname', 'AvailId', 'EndDate', 'StartDate', 
        'Start Year', 'End Year', 'Duration', 'Publishedto[0][0]', 
        'Publishedto[0][1]', 'Publishedto[1][0]', 'Publishedto[1][1]', 'AvailabilityId'
    ]
    df_ma = df_ma[cols_ma_order]

    # 2. BUILD MASTER DATA
    print("Processing Master Data sheet...")
    # Clean assignments ids and slotRequestIds
    df_assignments["slotRequestId"] = df_assignments["slotRequestId"].astype(str).str.strip()
    df_slots["_id"] = df_slots["_id"].astype(str).str.strip()

    # Assign order index to assignments to preserve original order after merge
    df_assignments["ass_order_idx"] = range(len(df_assignments))

    # Left join slots to assignments
    df_joined_raw = df_slots.merge(df_assignments, left_on="_id", right_on="slotRequestId", how="left", suffixes=("_slot", "_assign"))
    
    # Sort: all slots with assignments first (ordered by assignments order), then slots without assignments
    df_with_ass = df_joined_raw[df_joined_raw["_id_assign"].notna()].sort_values("ass_order_idx")
    df_no_ass = df_joined_raw[df_joined_raw["_id_assign"].isna()]
    df_joined = pd.concat([df_with_ass, df_no_ass], ignore_index=True)
    
    # Pre-compute assignment count per slot for use in Confirmation Status logic
    slot_assignment_counts = df_assignments.groupby("slotRequestId").size().to_dict()
    
    # Identify slots that have a "placeholder" assignment (no assigneeId, null status)
    placeholder_mask = (
        df_assignments["assigneeId"].isna() &
        (df_assignments["status"].isna() | df_assignments["status"].isin(["", "nan", "None"]))
    )
    slots_with_placeholder = set(df_assignments[placeholder_mask]["slotRequestId"].dropna().astype(str))

    print(f"Joined and ordered raw shape: {df_joined.shape}")

    tenant_id = str(df_tenant["_id"].iloc[0]).strip() if not df_tenant.empty else ""
    tenant_name = str(df_tenant["name"].iloc[0]).strip() if not df_tenant.empty else ""
    report_records = []
    report_date = datetime.datetime(2026, 5, 29)  # Stored in 'Current date' column + used for age calc

    # Pre-parse common date columns to optimize loop speed
    df_joined["created_dt"] = pd.to_datetime(df_joined["_createdAt_slot"]).dt.tz_localize(None)
    df_joined["updated_dt"] = pd.to_datetime(df_joined["_updatedAt_slot"]).dt.tz_localize(None)
    df_joined["approved_dt"] = pd.to_datetime(df_joined["approvedAt"]).dt.tz_localize(None)
    df_joined["start_dt"] = pd.to_datetime(df_joined["startDate_slot"]).dt.tz_localize(None)
    df_joined["end_dt"] = pd.to_datetime(df_joined["endDate_slot"]).dt.tz_localize(None)
    df_joined["placement_start_dt"] = pd.to_datetime(df_joined["startDate_assign"]).dt.tz_localize(None)
    df_joined["placement_end_dt"] = pd.to_datetime(df_joined["endDate_assign"]).dt.tz_localize(None)
    df_joined["assignee_created_dt"] = pd.to_datetime(df_joined["_createdAt_assign"]).dt.tz_localize(None)

    seen_slots = set()
    for idx, row in df_joined.iterrows():
        # Core fields
        slot_id = str(row["_id_slot"])
        is_duplicate = slot_id in seen_slots
        seen_slots.add(slot_id)
        
        ass_id = str(row["_id_assign"]).strip() if pd.notna(row["_id_assign"]) else ""
        has_assignment = bool(ass_id)
        
        # 1. Discipline IDs & Names
        disp_id = ""
        # iterate slot request columns to find first discipline ID in curriculum
        for col_i in range(11):
            c_col = f"curriculum[{col_i}].disciplineId_slot"
            if c_col in row and pd.notna(row[c_col]):
                disp_id = str(row[c_col]).strip()
                break
        
        disp_name = discipline_dict.get(disp_id, "")

        # 2. Age - In days & Ageing
        created_dt = row["created_dt"]
        updated_dt = row["updated_dt"]
        approved_dt = row["approved_dt"]
        status_slot = str(row["status_slot"]).strip()
        
        age = np.nan
        if pd.notna(approved_dt):
            age = (approved_dt - created_dt).days
        elif status_slot in ["In-Progress", "In Progress", "Pending", "Draft"]:
            age = (report_date - created_dt).days
        else:
            if pd.notna(updated_dt):
                age = (updated_dt - created_dt).days
            else:
                age = 0
                
        ageing = ""
        if pd.notna(age):
            age = int(age)
            if age <= 7:
                ageing = "<= 7 Days"
            elif age <= 15:
                ageing = "<= 15 Days"
            elif age <= 30:
                ageing = "<= 30 Days"
            else:
                ageing = "> 30 Days"

        # 3. Availability details
        avail_id = str(row["availabilityId_slot"]).strip() if pd.notna(row["availabilityId_slot"]) else ""
        avail_info = avail_lookup.get(avail_id, {})
        
        avail_start_dt = to_date_only(avail_info.get("StartDate"))
        avail_end_dt = to_date_only(avail_info.get("EndDate"))
        
        avail_start_year = avail_start_dt.strftime("%Y") if avail_start_dt else ""
        avail_end_year = avail_end_dt.strftime("%Y") if avail_end_dt else ""
        
        avail_duration = ""
        if avail_start_year and avail_end_year:
            avail_duration = f"{avail_start_year} - {avail_end_year}"
            
        avail_name = str(avail_info.get("availabilityname", "")).strip() if avail_info else ""

        # 4. School details
        school_id = str(row["oneSchoolId_slot"]).strip() if pd.notna(row["oneSchoolId_slot"]) else ""
        school_info = school_dict.get(school_id, {})
        school_name = school_info.get("name", "") if school_info else ""
        
        concat_name = f"H2 Health {school_name}".strip()
        aff_type = "Affiliated" if school_id in partner_school_ids else "Not Affiliated"
        
        prism_school = school_info.get("prismTenantId") if school_info else ""
        if pd.isna(prism_school) or not prism_school:
            prism_school = school_id
        prism_flag = is_prism_school(prism_school)

        # 5. Slots counts
        req_slots = row["requestedSlots"]
        app_slots = row["approvedSlots"]
        action_by = row["actionBy"]

        # 6. Graduation & Rotations
        grad_year = row["graduation.date.year_slot"]
        grad_month = row["graduation.date.month_slot"]
        
        grad_month_name = get_month_name(grad_month)
        
        grad_concat = None
        if pd.notna(grad_year) and pd.notna(grad_month):
            try:
                grad_concat = datetime.datetime(int(float(grad_year)), int(float(grad_month)), 1)
            except:
                grad_concat = f"{grad_month_name} {int(float(grad_year))}"
                
        rot_start_dt = to_date_only(row["startDate_slot"])
        rot_end_dt = to_date_only(row["endDate_slot"])
        
        rot_start_year = int(rot_start_dt.year) if rot_start_dt else None
        rot_end_year = int(rot_end_dt.year) if rot_end_dt else None
        
        rot_start_month = month_names.get(rot_start_dt.month, "") if rot_start_dt else ""
        rot_end_month = month_names.get(rot_end_dt.month, "") if rot_end_dt else ""
        
        rot_concat = None
        if rot_start_dt:
            rot_concat = datetime.datetime(rot_start_dt.year, rot_start_dt.month, 1)

        # 7. Student Year
        student_year = ""
        if pd.notna(grad_year) and rot_start_year is not None:
            try:
                diff = int(float(grad_year)) - rot_start_year
                if diff < 0:
                    student_year = "Graduated"
                elif diff == 0:
                    student_year = "Graduation Year"
                elif diff == 1:
                    student_year = "1 Year to Graduation"
                else:
                    student_year = f"{diff} Years to Graduation"
            except:
                pass

        # 8. Placement & requested start/end
        placement_start = to_date_only(row["startDate_assign"]) if has_assignment else None
        placement_end = to_date_only(row["endDate_assign"]) if has_assignment else None
        
        requested_start = to_date_only(row["startDate_slot"])
        requested_end = to_date_only(row["endDate_slot"])

        # 9. Assignment details
        compliant_status = np.nan
        if has_assignment:
            compliant_val = row.get("caas.compliant")
            if pd.isna(compliant_val):
                compliant_status = "Unknown"
            elif isinstance(compliant_val, bool):
                compliant_status = "Compliant" if compliant_val else "Not Compliant"
            elif str(compliant_val).strip().lower() == "true":
                compliant_status = "Compliant"
            elif str(compliant_val).strip().lower() == "false":
                compliant_status = "Not Compliant"
            else:
                compliant_status = "Unknown"

        if has_assignment:
            status_assign = row["status_assign"]
            raw_status = str(status_assign).strip() if pd.notna(status_assign) else ""
            if not raw_status or raw_status.lower() in ["nan", "none"]:
                if slot_id in slots_with_placeholder:
                    conf_status = "To Be Scheduled"
                else:
                    std_can_view = row.get("policy.stdCanViewAsnm")
                    can_view = (std_can_view == True or str(std_can_view).strip().lower() in ["true", "1", "t"])
                    conf_status = "Confirmed" if can_view else "Not Confirmed"
            elif raw_status.lower() == "active":
                conf_status = "Confirmed"
            else:
                conf_status = raw_status
        else:
            status_slot_val = str(row["status_slot"]).strip() if pd.notna(row["status_slot"]) else ""
            if status_slot_val.lower() == "approved":
                conf_status = "To Be Scheduled"
            elif status_slot_val.lower() == "draft":
                conf_status = "In-Progress"
            else:
                conf_status = status_slot_val
        group_id = str(row["groupId"]).strip() if pd.notna(row["groupId"]) else ""
        updated_by = str(row["updatedBy_assign"]).strip() if has_assignment and pd.notna(row["updatedBy_assign"]) else None
        
        # Location Lookup
        loc_id = str(row["locationId"]).strip() if pd.notna(row["locationId"]) else ""
        loc_info = location_dict.get(loc_id, {})
        loc_name = loc_info.get("name", "") if loc_info else ""
        loc_state = loc_info.get("addresses[0].state", "") if loc_info else ""
        
        loc_group_id = loc_info.get("groups[0]") if loc_info else ""
        loc_group_name = lookup_group_dict.get(loc_group_id, "") if pd.notna(loc_group_id) else ""

        assignee_id = str(row["assigneeId"]).strip() if pd.notna(row["assigneeId"]) else ""
        ast_info = assignee_dict.get(assignee_id, {})
        assignee_created = to_date_only(ast_info.get("_createdAt")) if has_assignment and ast_info else None
        student_name = format_name(ast_info.get("firstName"), ast_info.get("lastName")) if ast_info else ""
        student_email = format_email(ast_info.get("userEmail")) if ast_info else ""
        
        if assignee_id:
            scheduled_status = "Scheduled"
        elif has_assignment:
            scheduled_status = "Unscheduled"
        else:
            scheduled_status = ""
        
        display_id_ass = row["displayId_assign"]
        student_view_ass = row["policy.stdCanViewAsnm"] if has_assignment else ""
        pay_status = str(row["policy.subscription.payStatus"]).strip() if pd.notna(row["policy.subscription.payStatus"]) else ""
        if pay_status:
            pay_status = pay_status.title()
            
        caas_email = format_email(row.get("caas.userEmail"))

        # Preceptor details
        preceptor_id = str(row["preceptors[0]"]).strip() if pd.notna(row["preceptors[0]"]) else ""
        prep_info = profile_dict.get(preceptor_id, {})
        
        if not has_assignment:
            preceptor_name = ""
            preceptor_email = ""
        elif prep_info:
            preceptor_name = format_name(prep_info.get("firstName"), prep_info.get("lastName"))
            preceptor_email = format_email(prep_info.get("userEmail"))
            if not preceptor_name:
                preceptor_name = "NA"
            if not preceptor_email:
                preceptor_email = "NA"
        else:
            preceptor_name = "NA"
            preceptor_email = "NA"

        ass_status = str(row["status_assign"]).strip() if pd.notna(row["status_assign"]) else ""
        group_name_col = row["groupName"] if has_assignment else np.nan
        cancelled_note = str(row["notes[0]"]).strip() if pd.notna(row["notes[0]"]) else ""

        report_row = {
            "_id": slot_id,
            "IDs": disp_id,
            "Names": disp_name,
            "Status": status_slot,
            "Age - In days": age,
            "Ageing": ageing,
            "Current date": report_date,
            "_createdAt": to_date_only(row["_createdAt_slot"]),
            "_updatedAt": to_date_only(row["_updatedAt_slot"]),
            "approvedAt": to_date_only(row["approvedAt"]),
            "createdBy": str(row["createdBy_slot"]).strip() if pd.notna(row["createdBy_slot"]) else "",
            "Availability Start Year": float(avail_start_year) if avail_start_year else np.nan,
            "Availability End Year": float(avail_end_year) if avail_end_year else np.nan,
            "Availability Duration": avail_duration,
            "Availability[0].startDate": avail_start_dt,
            "Availability[0].endDate": avail_end_dt,
            "displayId": row["displayId_slot"],
            "SiteID": tenant_id,
            "Site_Name": tenant_name,
            "Availability_Id": avail_id,
            "Availability_Name": avail_name,
            "SchoolId": school_id,
            "School": school_name,
            "Concat": concat_name,
            "Aff type": aff_type,
            "PrismSchool[0]": prism_school,
            "Prism -School Yes or No": prism_flag,
            "RequestedSlots": float(req_slots) if pd.notna(req_slots) and not is_duplicate else np.nan,
            "ApprovedSlots": float(app_slots) if pd.notna(app_slots) and not is_duplicate else (0.0 if not is_duplicate and not has_assignment else np.nan),
            "actionBy": action_by,
            "Graduation.date.year": float(grad_year) if pd.notna(grad_year) else np.nan,
            "Graduation.date.month": float(grad_month) if pd.notna(grad_month) else np.nan,
            "Graduation Month": grad_month_name,
            "Rotation start Year": float(rot_start_year) if rot_start_year is not None else np.nan,
            "Rotation End Year": float(rot_end_year) if rot_end_year is not None else np.nan,
            "Rotation start Month": rot_start_month,
            "Rotation End Month": rot_end_month,
            "Graduation Concat": grad_concat,
            "Rotation Concat": rot_concat,
            "Student Year": student_year,
            "PlacementStartDate": placement_start,
            "PlacementEndDate": placement_end,
            "requestedStartdate": requested_start,
            "requestedEndDate": requested_end,
            "Req Year": "",
            "Req Start Month ": "",
            "Compliant Status": compliant_status,
            "Confirmation Status": conf_status,
            "GroupId": group_id,
            "updatedBy": updated_by,
            "location": loc_id,
            "LocationName[0]": loc_name,
            "Location State": loc_state,
            "Location group Names": loc_group_name,
            "AssigneId": assignee_id,
            "Assignee CreatedAt": assignee_created,
            "StudentName": student_name,
            "Email": student_email,
            "Scheduled Status": scheduled_status,
            "displayIdAssignement": float(display_id_ass) if pd.notna(display_id_ass) else np.nan,
            "AssignmentId": ass_id,
            "StudentViewAssignment": student_view_ass,
            "PayStatus": pay_status,
            "caas.userEmail": caas_email,
            "Preceptor Full Name": preceptor_name,
            "Preceptor_Email": preceptor_email,
            "Assignment Status": ass_status,
            "GroupName[0]": group_name_col,
            "CancelledNote": cancelled_note,
            "Graduation": grad_concat
        }
        report_records.append(report_row)

    df_master = pd.DataFrame(report_records)

    # Reorder columns to exactly match master
    cols_master_order = [
        '_id', 'IDs', 'Names', 'Status', 'Age - In days', 'Ageing', 'Current date', 
        '_createdAt', '_updatedAt', 'approvedAt', 'createdBy', 'Availability Start Year', 
        'Availability End Year', 'Availability Duration', 'Availability[0].startDate', 
        'Availability[0].endDate', 'displayId', 'SiteID', 'Site_Name', 'Availability_Id', 
        'Availability_Name', 'SchoolId', 'School', 'Concat', 'Aff type', 'PrismSchool[0]', 
        'Prism -School Yes or No', 'RequestedSlots', 'ApprovedSlots', 'actionBy', 
        'Graduation.date.year', 'Graduation.date.month', 'Graduation Month', 
        'Rotation start Year', 'Rotation End Year', 'Rotation start Month', 
        'Rotation End Month', 'Graduation Concat', 'Rotation Concat', 'Student Year', 
        'PlacementStartDate', 'PlacementEndDate', 'requestedStartdate', 'requestedEndDate', 
        'Req Year', 'Req Start Month ', 'Compliant Status', 'Confirmation Status', 
        'GroupId', 'updatedBy', 'location', 'LocationName[0]', 'Location State', 
        'Location group Names', 'AssigneId', 'Assignee CreatedAt', 'StudentName', 'Email', 
        'Scheduled Status', 'displayIdAssignement', 'AssignmentId', 'StudentViewAssignment', 
        'PayStatus', 'caas.userEmail', 'Preceptor Full Name', 'Preceptor_Email', 
        'Assignment Status', 'GroupName[0]', 'CancelledNote', 'Graduation'
    ]
    df_master = df_master[cols_master_order]

    # Format the current date dynamically (e.g. 1st June 2026)
    def get_ordinal_suffix(day):
        if 11 <= day <= 13:
            return "th"
        return {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")

    now = datetime.datetime.now()
    suffix = get_ordinal_suffix(now.day)
    current_date_str = f"{now.day}{suffix} {now.strftime('%B %Y')}"

    # Save to copying template file
    orig_report_path = os.path.join(workspace_dir, "template.xlsx")
    if not os.path.exists(orig_report_path):
        orig_report_path = os.path.join(workspace_dir, "H2 Health Report 29th May 2026.xlsx")
    out_report_path = os.path.join(workspace_dir, f"{tenant_name} Report {current_date_str}.xlsx")

    # Copy original to a temp working copy
    temp_dir = os.environ.get("TEMP", workspace_dir)
    temp_work_path = os.path.join(temp_dir, "H2_Health_Report_work.xlsx")
    print(f"Creating a working copy from template...")
    shutil.copyfile(orig_report_path, temp_work_path)

    print("Opening working copy using openpyxl...")
    wb = openpyxl.load_workbook(temp_work_path, data_only=False)

    # 1. Update Master Avail 29th may
    print("Writing Master Avail 29th may sheet...")
    ws_avail = wb["Master Avail 29th may"]
    
    # Clear existing rows (leaving headers)
    print(f"Clearing old rows in Master Avail (total rows: {ws_avail.max_row})...")
    for r in range(2, ws_avail.max_row + 1):
        for c in range(1, ws_avail.max_column + 1):
            ws_avail.cell(row=r, column=c).value = None

    for r_idx, row_values in enumerate(df_ma.values, 2):
        for c_idx, val in enumerate(row_values, 1):
            ws_avail.cell(row=r_idx, column=c_idx).value = val

    # 2. Update Master Data 29th May
    print("Writing Master Data 29th May sheet...")
    ws_data = wb["Master Data 29th May"]
    
    print(f"Clearing old rows in Master Data (total rows: {ws_data.max_row})...")
    for r in range(2, ws_data.max_row + 1):
        for c in range(1, ws_data.max_column + 1):
            ws_data.cell(row=r, column=c).value = None

    for r_idx, row_values in enumerate(df_master.values, 2):
        for c_idx, val in enumerate(row_values, 1):
            col_name = cols_master_order[c_idx - 1]
            if col_name == "Req Year":
                ws_data.cell(row=r_idx, column=c_idx).value = f'=TEXT(AQ{r_idx},"YYYY")'
            elif col_name == "Req Start Month ":
                ws_data.cell(row=r_idx, column=c_idx).value = f'=TEXT(AQ{r_idx},"MMM")'
            else:
                if pd.isna(val):
                    ws_data.cell(row=r_idx, column=c_idx).value = None
                else:
                    ws_data.cell(row=r_idx, column=c_idx).value = val

    print("Saving the modified workbook back to temp location...")
    wb.save(temp_work_path)
    wb.close()
    
    print("Copying the saved workbook to the final destination...")
    
    import time
    saved_path = None
    for attempt in range(3):
        try:
            shutil.copyfile(temp_work_path, out_report_path)
            saved_path = out_report_path
            break
        except PermissionError:
            if attempt < 2:
                print(f"  Output file locked, retrying in 3s... (attempt {attempt+1}/3)")
                time.sleep(3)
    
    if saved_path is None:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = os.path.join(workspace_dir, f"{tenant_name} Report {current_date_str}_output_{ts}.xlsx")
        shutil.copyfile(temp_work_path, fallback_path)
        saved_path = fallback_path
    
    try:
        os.remove(temp_work_path)
    except:
        pass
    print(f"Workbook successfully saved to: {saved_path}")

    # 3. GENERATE CLINICAL ASSIGNMENTS REPORT (from generate_report.py)
    print("\nProcessing Clinical Assignments Report...")
    # Index lookups for performance and correctness
    assignee_dict = df_assignees.set_index("_id").to_dict(orient="index")
    school_dict = df_schools.set_index("_id").to_dict(orient="index")
    location_dict = df_locations.set_index("_id").to_dict(orient="index")
    profile_dict = df_profiles.set_index("_id").to_dict(orient="index")
    discipline_dict = df_disciplines.set_index("_id").to_dict(orient="index")
    availability_dict = df_avail.set_index("_id").to_dict(orient="index")

    # Match student profile phones (profiles mapped by oneProfileId)
    student_profile_phone_dict = {}
    for pid, p_row in df_profiles.iterrows():
        one_prof_id = p_row.get("oneProfileId")
        if pd.notna(one_prof_id):
            phone = p_row.get("addresses[0].contacts[0].phone")
            if pd.notna(phone):
                student_profile_phone_dict[str(one_prof_id)] = phone

    report_records = []

    # Identify discipline and specialization columns
    disp_cols = [c for c in df_assignments.columns if 'disciplineId' in c]
    spec_cols = [c for c in df_assignments.columns if 'specializationId' in c]

    for idx, r_row in df_assignments.iterrows():
        # 1. Student Details (Assignee)
        assignee_id = str(r_row.get("assigneeId"))
        student_name = ""
        student_email = ""
        student_phone = ""
        
        if assignee_id in assignee_dict:
            ast_row = assignee_dict[assignee_id]
            student_name = format_name_report(ast_row.get("firstName"), ast_row.get("lastName"))
            student_email = str(ast_row.get("userEmail")).strip() if pd.notna(ast_row.get("userEmail")) else ""
            
            # Lookup Student Phone via oneProfileId
            one_prof_id = str(ast_row.get("oneProfileId")) if pd.notna(ast_row.get("oneProfileId")) else ""
            if one_prof_id in student_profile_phone_dict:
                student_phone = format_phone_report(student_profile_phone_dict[one_prof_id])

        # 2. School Name
        school_id = str(r_row.get("oneSchoolId"))
        school_name = ""
        if school_id in school_dict:
            school_name = school_dict[school_id].get("name", "")

        # 3. Location Details
        loc_id = str(r_row.get("locationId"))
        location_name = ""
        city = ""
        state = ""
        zip_code = ""
        if loc_id in location_dict:
            loc_row = location_dict[loc_id]
            location_name = loc_row.get("name", "")
            city = loc_row.get("addresses[0].city", "")
            state = loc_row.get("addresses[0].state", "")
            zip_code = loc_row.get("addresses[0].zip", "")
            if pd.notna(zip_code):
                if isinstance(zip_code, float) and zip_code.is_integer():
                    zip_code = str(int(zip_code))
                else:
                    zip_code = str(zip_code).strip()

        # 4. Discipline
        discipline_name = ""
        for col in disp_cols:
            d_id = str(r_row.get(col))
            if d_id in discipline_dict:
                discipline_name = discipline_dict[d_id].get("name", "")
                break # Use first found

        # 5. Specialty (Specialization)
        specialty_name = ""
        for col in spec_cols:
            s_id = str(r_row.get(col))
            if s_id in discipline_dict:
                specialty_name = discipline_dict[s_id].get("name", "")
                break
        
        # Fallback/heuristic for specialty from availability name if empty
        av_name = str(r_row.get("availabilities[0].name")) if pd.notna(r_row.get("availabilities[0].name")) else ""
        if not specialty_name and av_name:
            if "OP Pelvic PT" in av_name or "Pelvic health" in av_name:
                specialty_name = "Pelvic Health PT"
            elif "Pediatric" in av_name or "Peds" in av_name:
                specialty_name = "Pediatric PT"
            elif "ATC" in av_name or "Davis High" in av_name:
                specialty_name = "Athletic Training"
            elif "OP Ortho" in av_name or "Ortho" in av_name:
                specialty_name = "Orthopedic PT"

        # 6. Dates
        start_date = format_date_report(r_row.get("startDate"))
        end_date = format_date_report(r_row.get("endDate"))

        # 7. Preceptor Details
        preceptor_id = str(r_row.get("preceptors[0]"))
        preceptor_name = ""
        preceptor_email = ""
        preceptor_phone = ""
        if preceptor_id in profile_dict:
            prep_row = profile_dict[preceptor_id]
            preceptor_name = format_name_report(prep_row.get("firstName"), prep_row.get("lastName"))
            preceptor_email = str(prep_row.get("userEmail")).strip() if pd.notna(prep_row.get("userEmail")) else ""
            preceptor_phone = format_phone_report(prep_row.get("addresses[0].contacts[0].phone"))

        # 8. Status
        status = str(r_row.get("status")).strip() if pd.notna(r_row.get("status")) else ""
        if status:
            status = status.title()

        # 9. Availability Name
        if not av_name:
            av_id = str(r_row.get("availabilityId"))
            if av_id in availability_dict:
                av_name = availability_dict[av_id].get("name", "")

        # 10. Pay Status
        pay_status = str(r_row.get("policy.subscription.payStatus")).strip() if pd.notna(r_row.get("policy.subscription.payStatus")) else ""
        if pay_status:
            pay_status = pay_status.title()

        # 11. Rotation, Course & Code
        rotation_name = str(r_row.get("ext.rotationDetails.name")).strip() if pd.notna(r_row.get("ext.rotationDetails.name")) else ""
        course_name = str(r_row.get("ext.courseDetails.name")).strip() if pd.notna(r_row.get("ext.courseDetails.name")) else ""
        course_code = str(r_row.get("ext.courseDetails.code")).strip() if pd.notna(r_row.get("ext.courseDetails.code")) else ""

        # Construct unified row
        report_row = {
            "Student Name": student_name,
            "School Name": school_name,
            "Email": student_email,
            "Phone": student_phone,
            "Discipline": discipline_name,
            "Specialty": specialty_name,
            "Location": location_name,
            "City": city,
            "State": state,
            "Zip": zip_code,
            "Start Date": start_date,
            "End Date": end_date,
            "Preceptor": preceptor_name,
            "Email ": preceptor_email,
            "Phone ": preceptor_phone,
            "Status": status,
            "Availability Name": av_name,
            "Pay Status": pay_status,
            "Rotation Name": rotation_name,
            "Course Name": course_name,
            "Code": course_code
        }
        report_records.append(report_row)

    # Convert to DataFrame
    report_df = pd.DataFrame(report_records)
    
    cols_order = [
        "Student Name", "School Name", "Email", "Phone",
        "Discipline", "Specialty", "Location", "City",
        "State", "Zip", "Start Date", "End Date",
        "Preceptor", "Email ", "Phone ", "Status",
        "Availability Name", "Pay Status", "Rotation Name",
        "Course Name", "Code"
    ]
    report_df = report_df[cols_order]

    # Save to CSV
    csv_path = os.path.join(workspace_dir, "master_report.csv")
    report_df.to_csv(csv_path, index=False, encoding="utf-8")
    print(f"Successfully saved CSV report to: {csv_path}")

    # Now execute the TAI report logic as integrated!
    try:
        generate_tai_report()
    except Exception as e:
        print(f"Error during TAI report generation: {e}")



if __name__ == "__main__":
    main()
